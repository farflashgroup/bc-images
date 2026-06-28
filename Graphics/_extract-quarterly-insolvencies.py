"""Build quarterly consumer insolvency series from OSB xlsx files."""
from __future__ import annotations

import json
import urllib.request
from io import BytesIO
from pathlib import Path

import openpyxl

URLS = {
    2011: "https://ised-isde.canada.ca/site/office-superintendent-bankruptcy/sites/default/files/attachments/2022/Insolvency_Statistiques_insolvabilite_2011.xlsx",
    2012: "https://www.ic.gc.ca/eic/site/bsf-osb.nsf/vwapj/Insolvency_Statistiques_insolvabilite_2012.xlsx/$file/Insolvency_Statistiques_insolvabilite_2012.xlsx",
    2013: "https://www.ic.gc.ca/eic/site/bsf-osb.nsf/vwapj/Insolvency_Statistiques_insolvabilite_2013.xlsx/$file/Insolvency_Statistiques_insolvabilite_2013.xlsx",
    2014: "https://www.ic.gc.ca/eic/site/bsf-osb.nsf/vwapj/Insolvency_Statistiques_insolvabilite_2014.xlsx/$file/Insolvency_Statistiques_insolvabilite_2014.xlsx",
    2015: "https://www.ic.gc.ca/eic/site/bsf-osb.nsf/vwapj/Insolvency_Statistiques_insolvabilite_2015.xlsx/$file/Insolvency_Statistiques_insolvabilite_2015.xlsx",
    2016: "https://www.ic.gc.ca/eic/site/bsf-osb.nsf/vwapj/Insolvency_Statistiques_insolvabilite_2016.xlsx/$file/Insolvency_Statistiques_insolvabilite_2016.xlsx",
    2017: "https://www.ic.gc.ca/eic/site/bsf-osb.nsf/vwapj/Insolvency_Statistiques_insolvabilite_2017.xlsx/$file/Insolvency_Statistiques_insolvabilite_2017.xlsx",
    2018: "https://www.ic.gc.ca/eic/site/bsf-osb.nsf/vwapj/Insolvency_Statistiques_insolvabilite_2018.xlsx/$file/Insolvency_Statistiques_insolvabilite_2018.xlsx",
    2019: "https://www.ic.gc.ca/eic/site/bsf-osb.nsf/vwapj/Insolvency_Statistiques_insolvabilite_2019.xlsx/$file/Insolvency_Statistiques_insolvabilite_2019.xlsx",
    2020: "https://www.ic.gc.ca/eic/site/bsf-osb.nsf/vwapj/Insolvency_Statistiques_insolvabilite_2020.xlsx/$file/Insolvency_Statistiques_insolvabilite_2020.xlsx",
    2021: "https://www.ic.gc.ca/eic/site/bsf-osb.nsf/vwapj/Insolvency_Statistiques_insolvabilite_2021.xlsx/$file/Insolvency_Statistiques_insolvabilite_2021.xlsx",
    2022: "https://ised-isde.canada.ca/site/office-superintendent-bankruptcy/sites/default/files/attachments/2023/Insolvency_Statistiques_insolvabilite_2022_.xlsx",
    2023: "https://ised-isde.canada.ca/site/office-superintendent-bankruptcy/sites/default/files/documents/insolvency_statistiques_insolvabilite_2023-annual.xlsx",
    2024: "https://ised-isde.canada.ca/site/office-superintendent-bankruptcy/sites/default/files/documents/insolvency_statistiques_insolvabilite_2024_annual.xlsx",
    2025: "https://ised-isde.canada.ca/site/office-superintendent-bankruptcy/sites/default/files/documents/insolvency_statistiques_insolvabilite_dec_2025.xlsx",
}

PROVINCE_KEYS = [
    "Newfoundland and Labrador",
    "Prince Edward Island",
    "Nova Scotia",
    "New Brunswick",
    "Quebec",
    "Ontario",
    "Manitoba",
    "Saskatchewan",
    "Alberta",
    "British Columbia",
    "Northwest Territories",
    "Yukon",
    "Nunavut",
]


def province_key(label) -> str | None:
    if not label:
        return None
    en = str(label).split("/")[0].strip().replace("New-Brunswick", "New Brunswick")
    if en.startswith("Yukon"):
        return "Yukon"
    for key in PROVINCE_KEYS:
        if en.startswith(key):
            return key
    return None

# OSB quarterly consumer insolvencies, Canada total (2009–2010 annual reports)
LEGACY = {
    "2009Q1": 38715,
    "2009Q2": 38446,
    "2009Q3": 36654,
    "2009Q4": 37897,
    "2010Q1": 37015,
    "2010Q2": 36786,
    "2010Q3": 33656,
    "2010Q4": 32777,
}


def fetch(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    return urllib.request.urlopen(req, timeout=90).read()


def quarter_values(row) -> list[int] | None:
    vals = []
    for v in row[1:5]:
        if isinstance(v, (int, float)) and v >= 0:
            vals.append(int(v))
        else:
            return None
    return vals if len(vals) == 4 else None


def extract_er_sheet(ws) -> list[int]:
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Canada":
            q = quarter_values(row)
            if q and q[0] > 5000:
                return q

    seen: set[str] = set()
    totals = [0, 0, 0, 0]
    for row in ws.iter_rows(values_only=True):
        key = province_key(row[0])
        if not key or key in seen:
            continue
        q = quarter_values(row)
        if not q:
            continue
        seen.add(key)
        for i in range(4):
            totals[i] += q[i]
    if len(seen) < 12:
        raise ValueError(f"Only found {len(seen)} provinces/territories")
    return totals


def extract_year(year: int) -> list[int]:
    wb = openpyxl.load_workbook(BytesIO(fetch(URLS[year])), read_only=True, data_only=True)
    sheet_name = next(n for n in wb.sheetnames if "ER_RE" in n)
    return extract_er_sheet(wb[sheet_name])


def main() -> None:
    series: list[dict] = []

    for key in sorted(LEGACY):
        y, q = key.split("Q")
        series.append({"year": int(y), "quarter": int(q), "value": LEGACY[key]})

    for year in range(2011, 2026):
        qvals = extract_year(year)
        for qi, value in enumerate(qvals, start=1):
            series.append({"year": year, "quarter": qi, "value": value})
        print(f"{year}: {qvals} ({sum(qvals)})")

    series.append({"year": 2026, "quarter": 1, "value": 37121, "latest": True})

    out = Path(__file__).parent / "article-assets" / "insolvency-quarterly-osb.json"
    out.write_text(json.dumps(series, indent=2), encoding="utf-8")
    print(f"\nWrote {len(series)} quarters -> {out.name}")
    peak = max(series, key=lambda d: d["value"])
    latest = series[-1]
    print(f"Peak: {peak['year']}Q{peak['quarter']} = {peak['value']}")
    print(f"Latest: {latest['year']}Q{latest['quarter']} = {latest['value']}")


if __name__ == "__main__":
    main()
